import time
import openpyxl
import matplotlib.pyplot as plt
from matplotlib import rc

# 한글 글꼴 설정
rc('font', family='Malgun Gothic')  # Windows 사용 시
plt.rcParams['axes.unicode_minus'] = False  # 음수 기호 깨짐 방지

# 엑셀 파일에서 데이터 읽기 함수
def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    labels = []
    values = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        labels.append(row[0])  # 후보자 이름
        values.append(row[1])  # 득표 수
    return labels, values

# 파이 그래프 그리기 함수 (반복 실행 포함)
def draw_pie_chart(file_path, interval=10):
    for _ in range(100):  # 반복 횟수만큼 실행
        labels, values = read_excel(file_path)  # 엑셀 데이터 읽기
        explode = [0.1 if value == max(values) else 0 for value in values]  # 최대 득표자 강조
        plt.pie(values, explode=explode, labels=labels, autopct='%1.1f%%', startangle=67)
        plt.title("실시간 득표율")
        plt.draw()  # 그래프 그리기
        plt.pause(1)  # 1초 대기
        plt.clf()     # 그래프 초기화
        time.sleep(interval)  # 지정된 간격 대기

# 메인 실행 함수
def main():
    file_path = r"C:\GitCode\CodeFolder\UNIV\VisualProgramming\HW4\elec.xlsx"
    draw_pie_chart(file_path)  # 그래프 실행

if __name__ == "__main__":
    main()

